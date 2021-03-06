
# makes by paraclin code calculating file

import os
from collections import defaultdict
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
#from openpyxl.compat import range
from openpyxl.styles import Border, Side, colors
from flask import g
from poly.utils.files import get_name_tail
from poly.reestr.invoice.impex.exp_inv import get_mo_smo_name
from poly.reestr.invoice.impex import config

def data_source_init():
    g.qurs = g.qonn.cursor(cursor_factory=psycopg2.extras.NamedTupleCursor)

def data_source_get():
    g.qurs.execute(config.GET_USL)
    return g.qurs.fetchall()

def data_source_close():
    g.qurs.close()


def exp_usl(app: object, smo: int, month: str, year: str, inv_path: str) -> (int, str):
    
    typ=6 # pmu
    tpl= config.TYPE[typ-1][2]
    sh1 = 'Лист1'
    xtpl = f'{tpl}.xlsx'
    xlr = os.path.join(inv_path, 'tpl', xtpl)   
    xout = f'{tpl}_{smo}_{month}_{year}_{get_name_tail(4)}.xlsx'
    
    mon= int(month)
    #smo= int(insurer)
    
    xlw = os.path.join(inv_path, xout)   

    period = 'За %s %s года' % (app.config['MONTH'][mon-1], year)

    wb = load_workbook(filename = xlr)
    #wb.active
    sheet = wb[sh1]

    border = Border(
        left=Side(border_style='thin', color=colors.BLACK),
        right=Side(border_style='thin', color=colors.BLACK),
        top=Side(border_style='thin', color=colors.BLACK),
        bottom=Side(border_style='thin', color=colors.BLACK)
    )

    cnt = 21
    rc = 1

    data_source_init()
    mo_name, smo_name = get_mo_smo_name(app, smo, config)
    sheet['B15'].value = '%s   %s' % (period, smo_name)

    for row in data_source_get():

        data = ( row.code_usl, row.name,  row.tarif, row.kol_usl, row.kol_usl*row.tarif)
        for xrow in range(cnt, cnt+1):
            for xcol in range(1, 6):
                c = sheet.cell(column=xcol, row=xrow, value= data[ xcol-1 ])
                c.border = border
        cnt += 1  
        rc += 1
    
    for xcol in range(1, 6):
        val=''
        if xcol==1:
            val="ИТОГО"
        if xcol==4:
            d = f'D{cnt-1}'
            val=f'=SUM(D21:D{cnt-1})'
        if xcol==5:
            d = f'E{cnt-1}'
            val=f'=SUM(E21:E{cnt-1})'
        c = sheet.cell(column=xcol, row=cnt, value= val)
        c.border = border
    
    sheet.cell(column=2, row=cnt+4, value= "Генеральный директор")
    sheet.cell(column=3, row=cnt+4, value= "ДЕНИСОВА С. А.")
    sheet.cell(column=2, row=cnt+6, value= "Исполнитель   202-51-45   Чамбайшин Ю.А.")	

    wb.save(xlw)
    wb.close()

    data_source_close()

    return rc, xlw
